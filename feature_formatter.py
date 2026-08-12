"""
FEATURE 1: Rendered Medicines Report Formatter.

Accepts a raw "Rendered Medicines" export in *any* reasonable column
layout — mixed-up order, inconsistent naming (e.g. "Cellphone Number" vs
"Contact Number"), split or combined name fields, messy whitespace/date
formats, PINs read in as floats, etc. — auto-detects each needed field,
cleans it, and always produces the same fixed 22-column report:

    Patient Name, Last Name, First Name, Middle Name, Patient PIN,
    Patient Source, Consultation Date, Rendered Date, End Visit By,
    ICD10 Code, ICD10 Description, Medicine, Medicine Category,
    Qty Prescribed, Qty Dispensed, Cost, Price, Total Cost, Total Price,
    Contact Number, Address, Notes
"""
import io
import zipfile

import openpyxl
import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from utils import (
    clean_date,
    clean_number,
    clean_pin,
    clean_str,
    find_column,
    format_contact_number,
    sanitize_filename_part,
    split_name_fallback,
)

OUTPUT_COLUMNS = [
    "Patient Name",
    "Last Name",
    "First Name",
    "Middle Name",
    "Patient PIN",
    "Patient Source",
    "Consultation Date",
    "Rendered Date",
    "End Visit By",
    "ICD10 Code",
    "ICD10 Description",
    "Medicine",
    "Medicine Category",
    "Qty Prescribed",
    "Qty Dispensed",
    "Cost",
    "Price",
    "Total Cost",
    "Total Price",
    "Contact Number",
    "Address",
    "Notes",
]

# Columns that get merged across a patient's multiple medicine rows
# (i.e. patient-identity/contact fields, not per-medicine fields).
MERGE_COLS = [
    "Patient Name",
    "Last Name",
    "First Name",
    "Middle Name",
    "Patient PIN",
    "Patient Source",
    "Consultation Date",
    "Rendered Date",
    "End Visit By",
    "Contact Number",
    "Address",
    "Notes",
]

CENTERED_COLS = [
    "Consultation Date",
    "Rendered Date",
    "Patient PIN",
    "Contact Number",
    "ICD10 Code",
]


def _detect_columns(df):
    category_col = find_column(df, ["Medicine Category", "Category", "Drug Category"])
    return {
        "patient_name": find_column(df, ["Patient Name", "Full Name", "Patient Full Name"]),
        "last": find_column(df, ["Last Name"]),
        "first": find_column(df, ["First Name"]),
        "middle": find_column(df, ["Middle Name"]),
        "pin": find_column(df, ["Patient PIN", "PIN", "PatientPIN"]),
        "source": find_column(df, ["Patient Source", "Source"]),
        "consult_date": find_column(df, ["Consultation Date"]),
        "rendered_date": find_column(df, ["Rendered Date", "Date Rendered"]),
        "end_visit_by": find_column(df, ["End Visit By"]),
        "icd_code": find_column(df, ["ICD10 Code", "ICD 10 Code", "ICD-10 Code"]),
        "icd_desc": find_column(df, ["ICD10 Description", "ICD 10 Description", "ICD-10 Description"]),
        "medicine": find_column(df, ["Medicine", "Medicine Name", "Drug", "Drug Name"], exclude=[category_col]),
        "category": category_col,
        "qty_prescribed": find_column(df, ["Qty Prescribed", "Quantity Prescribed"]),
        "qty_dispensed": find_column(df, ["Qty Dispensed", "Quantity Dispensed"]),
        "cost": find_column(df, ["Cost", "Unit Cost"]),
        "price": find_column(df, ["Price", "Unit Price"]),
        "total_cost": find_column(df, ["Total Cost"]),
        "total_price": find_column(df, ["Total Price"]),
        "phone": find_column(df, ["Contact Number", "Cellphone Number", "Mobile Number", "Phone Number"]),
        "address": find_column(df, ["Full Address", "Address", "Complete Address", "Home Address"]),
        "street": find_column(df, ["Street Name"]),
        "barangay": find_column(df, ["Barangay"]),
        "municipality": find_column(df, ["Municipality", "City"]),
        "province": find_column(df, ["Province"]),
        "notes": find_column(df, ["Notes", "Remarks"]),
    }


def _get(row, col):
    return row.get(col) if col else None


def _normalize_to_output_schema(df, cols):
    """Builds a DataFrame with exactly OUTPUT_COLUMNS, cleaned, regardless
    of how messy/mixed-up the source file's columns were."""
    records = []
    for _, row in df.iterrows():
        last_val = clean_str(_get(row, cols["last"])) if cols["last"] else ""
        first_val = clean_str(_get(row, cols["first"])) if cols["first"] else ""
        middle_val = clean_str(_get(row, cols["middle"])) if cols["middle"] else ""
        raw_name = clean_str(_get(row, cols["patient_name"])) if cols["patient_name"] else ""

        # Decided per row (not per file) so a file that mixes both naming
        # styles across different rows still resolves each row correctly.
        if last_val and first_val:
            last, first, middle = last_val, first_val, middle_val
            patient_name = raw_name or " ".join(p for p in [first, middle, last] if p)
        elif raw_name:
            last, first, middle = split_name_fallback(raw_name)
            patient_name = raw_name
        else:
            last, first, middle = last_val, first_val, middle_val
            patient_name = " ".join(p for p in [first, middle, last] if p)

        if cols["address"]:
            address = clean_str(_get(row, cols["address"]))
        else:
            parts = [
                clean_str(_get(row, cols[k]))
                for k in ("street", "barangay", "municipality", "province")
            ]
            address = ", ".join(p for p in parts if p)

        qty_dispensed = clean_number(_get(row, cols["qty_dispensed"]))
        cost = clean_number(_get(row, cols["cost"]))
        price = clean_number(_get(row, cols["price"]))

        total_cost = clean_number(_get(row, cols["total_cost"]))
        if total_cost == "" and qty_dispensed != "" and cost != "":
            total_cost = round(qty_dispensed * cost, 2)

        total_price = clean_number(_get(row, cols["total_price"]))
        if total_price == "" and qty_dispensed != "" and price != "":
            total_price = round(qty_dispensed * price, 2)

        records.append(
            {
                "Patient Name": patient_name,
                "Last Name": last,
                "First Name": first,
                "Middle Name": middle,
                "Patient PIN": clean_pin(_get(row, cols["pin"])),
                "Patient Source": clean_str(_get(row, cols["source"])),
                "Consultation Date": clean_date(_get(row, cols["consult_date"])),
                "Rendered Date": clean_date(_get(row, cols["rendered_date"])),
                "End Visit By": clean_str(_get(row, cols["end_visit_by"])),
                "ICD10 Code": clean_str(_get(row, cols["icd_code"])).upper(),
                "ICD10 Description": clean_str(_get(row, cols["icd_desc"])),
                "Medicine": clean_str(_get(row, cols["medicine"])),
                "Medicine Category": clean_str(_get(row, cols["category"])),
                "Qty Prescribed": clean_number(_get(row, cols["qty_prescribed"])),
                "Qty Dispensed": qty_dispensed,
                "Cost": cost,
                "Price": price,
                "Total Cost": total_cost,
                "Total Price": total_price,
                "Contact Number": format_contact_number(_get(row, cols["phone"])),
                "Address": address,
                "Notes": clean_str(_get(row, cols["notes"])),
            }
        )

    return pd.DataFrame(records, columns=OUTPUT_COLUMNS)


def _mapping_summary(cols):
    def fmt(key, label):
        val = cols.get(key)
        return f"`{val}`" if val else "not found — left blank"

    name_line = (
        f"`{cols['last']}` + `{cols['first']}`" + (f" + `{cols['middle']}`" if cols["middle"] else "")
        if cols["last"] and cols["first"]
        else (f"split from `{cols['patient_name']}`" if cols["patient_name"] else "not found — left blank")
    )
    address_line = (
        f"`{cols['address']}`"
        if cols["address"]
        else (
            "built from " + " + ".join(f"`{cols[k]}`" for k in ("street", "barangay", "municipality", "province") if cols[k])
            if any(cols[k] for k in ("street", "barangay", "municipality", "province"))
            else "not found — left blank"
        )
    )
    return f"""
- **Name** → {name_line}
- **Patient PIN** → {fmt('pin', 'PIN')}
- **Patient Source** → {fmt('source', 'Source')}
- **Consultation Date** → {fmt('consult_date', 'Consultation Date')}
- **Rendered Date** → {fmt('rendered_date', 'Rendered Date')}
- **ICD10 Code / Description** → {fmt('icd_code', '')} / {fmt('icd_desc', '')}
- **Medicine / Category** → {fmt('medicine', '')} / {fmt('category', '')}
- **Qty Prescribed / Dispensed** → {fmt('qty_prescribed', '')} / {fmt('qty_dispensed', '')}
- **Cost / Price / Total Cost / Total Price** → {fmt('cost','')} / {fmt('price','')} / {fmt('total_cost','')} / {fmt('total_price','')}
- **Contact Number** → {fmt('phone', '')}
- **Address** → {address_line}
- **Notes** → {fmt('notes', '')}
    """


def _sort_by_patient_source(df):
    """
    Sorts rows alphabetically by Patient Source (then Last Name, then First
    Name as tiebreakers), keeping every patient's own rows (e.g. multiple
    medicines) contiguous — required for the patient-block merging further
    down. Columns are already canonical at this point (post-normalization).
    """
    df = df.copy()
    df["_sort_source"] = df["Patient Source"].astype(str).str.strip().str.lower()
    df["_sort_last"] = df["Last Name"].astype(str).str.strip().str.lower()
    df["_sort_first"] = df["First Name"].astype(str).str.strip().str.lower()
    df["_sort_pin"] = df["Patient PIN"].astype(str)
    df = df.sort_values(
        by=["_sort_source", "_sort_last", "_sort_first", "_sort_pin"],
        kind="stable",
    ).reset_index(drop=True)
    return df.drop(columns=["_sort_source", "_sort_last", "_sort_first", "_sort_pin"])


def process_rendered_medicines(df_raw):
    """
    Normalizes the raw upload, then splits it into one workbook per unique
    Patient Source. Returns a list of (source_label, row_count, BytesIO)
    tuples, in the same alphabetical order as the sort applied to the data.
    """
    cols = _detect_columns(df_raw)

    with st.expander("Detected column mapping (click to verify)"):
        st.markdown(_mapping_summary(cols))

    df = _normalize_to_output_schema(df_raw, cols)
    df = _sort_by_patient_source(df)

    reports = []
    for source_val, group_df in df.groupby("Patient Source", sort=False):
        group_df = group_df.reset_index(drop=True)
        label = source_val if source_val else "Unspecified Source"
        buffer = _build_workbook(group_df)
        reports.append((label, len(group_df), buffer))

    return reports


def _build_workbook(df):
    """Builds one formatted Excel workbook (merged patient blocks, totals
    row, currency/qty formatting) from an already-normalized, single-source
    DataFrame with exactly OUTPUT_COLUMNS."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rendered Medicines"
    ws.views.sheetView[0].showGridLines = True

    headers = list(df.columns)

    header_fill = PatternFill(
        start_color="1F497D", end_color="1F497D", fill_type="solid"
    )
    header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    data_font = Font(name="Segoe UI", size=9.5)
    total_font = Font(name="Segoe UI", size=10, bold=True)

    fill_even = PatternFill(
        start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
    )
    fill_odd = PatternFill(
        start_color="F9FAFB", end_color="F9FAFB", fill_type="solid"
    )

    thin_border_side = Side(border_style="thin", color="D9D9D9")
    thin_border = Border(
        left=thin_border_side,
        right=thin_border_side,
        top=thin_border_side,
        bottom=thin_border_side,
    )
    top_thick_bottom_double = Border(
        top=Side(border_style="thin", color="1F497D"),
        bottom=Side(border_style="double", color="1F497D"),
    )

    align_header = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )
    align_right = Alignment(horizontal="right", vertical="center")
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_center_top = Alignment(
        horizontal="center", vertical="top", wrap_text=True
    )
    align_left_top = Alignment(
        horizontal="left", vertical="top", wrap_text=True
    )

    ws.row_dimensions[1].height = 28
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_header
        cell.border = thin_border

    patient_blocks = []
    current_pin = None
    block_start = 2

    for row_idx, row in df.iterrows():
        pin = row["Patient PIN"] if row["Patient PIN"] else f"NO_PIN_{row_idx}"
        excel_row = row_idx + 2
        if current_pin is None:
            current_pin = pin
            block_start = excel_row
        elif pin != current_pin:
            patient_blocks.append((block_start, excel_row - 1))
            current_pin = pin
            block_start = excel_row

    if block_start <= len(df) + 1:
        patient_blocks.append((block_start, len(df) + 1))

    for row_idx, row_data in enumerate(df.values, start=2):
        ws.row_dimensions[row_idx].height = 20
        block_idx = next(
            i for i, (s, e) in enumerate(patient_blocks) if s <= row_idx <= e
        )
        row_fill = fill_even if block_idx % 2 == 0 else fill_odd

        for col_idx, val in enumerate(row_data, start=1):
            col_name = headers[col_idx - 1]
            cell = ws.cell(row=row_idx, column=col_idx)

            cell.value = "" if pd.isna(val) else val
            cell.font = data_font
            cell.fill = row_fill
            cell.border = thin_border

            if col_name in ["Cost", "Price", "Total Cost", "Total Price"]:
                cell.number_format = "₱#,##0.00"
                cell.alignment = align_right
            elif col_name in ["Qty Prescribed", "Qty Dispensed"]:
                cell.number_format = "#,##0"
                cell.alignment = align_right
            elif col_name in CENTERED_COLS:
                cell.alignment = align_center
                if col_name in ["Patient PIN", "Contact Number"]:
                    cell.number_format = "@"
            else:
                cell.alignment = align_left

    for col_name in MERGE_COLS:
        if col_name in headers:
            col_idx = headers.index(col_name) + 1
            for start_row, end_row in patient_blocks:
                if start_row < end_row:
                    sub_vals = [
                        ws.cell(r, col_idx).value
                        for r in range(start_row, end_row + 1)
                    ]
                    if len(set(sub_vals)) == 1:
                        ws.merge_cells(
                            start_row=start_row,
                            start_column=col_idx,
                            end_row=end_row,
                            end_column=col_idx,
                        )
                        is_centered = col_name in CENTERED_COLS
                        target_align = (
                            align_center_top if is_centered else align_left_top
                        )

                        for r in range(start_row, end_row + 1):
                            c = ws.cell(row=r, column=col_idx)
                            c.alignment = target_align
                            c.border = thin_border

    total_row_idx = len(df) + 2
    ws.row_dimensions[total_row_idx].height = 24

    total_label = ws.cell(row=total_row_idx, column=1, value="Total")
    total_label.font = total_font
    total_label.alignment = Alignment(horizontal="left", vertical="center")

    for col_idx, col_name in enumerate(headers, start=1):
        cell = ws.cell(row=total_row_idx, column=col_idx)
        cell.font = total_font
        cell.border = top_thick_bottom_double
        col_letter = get_column_letter(col_idx)

        if col_name in ["Qty Prescribed", "Qty Dispensed"]:
            cell.value = f"=SUM({col_letter}2:{col_letter}{total_row_idx-1})"
            cell.number_format = "#,##0"
            cell.alignment = align_right
        elif col_name in ["Total Cost", "Total Price"]:
            cell.value = f"=SUM({col_letter}2:{col_letter}{total_row_idx-1})"
            cell.number_format = "₱#,##0.00"
            cell.alignment = align_right

    for col in ws.columns:
        col_idx = col[0].column
        col_name = headers[col_idx - 1]
        max_len = max(
            len(str(cell.value or "")) for cell in col if cell.value is not None
        )
        col_letter = get_column_letter(col_idx)

        if col_name in ["ICD10 Description", "Medicine"]:
            width = min(max(max_len + 3, 25), 45)
        elif col_name in ["Patient Name", "End Visit By"]:
            width = min(max(max_len + 3, 20), 30)
        elif col_name in ["Total Cost", "Total Price", "Cost", "Price"]:
            width = 14
        elif col_name in ["Qty Prescribed", "Qty Dispensed"]:
            width = 15
        else:
            width = min(max(max_len + 4, 12), 25)

        ws.column_dimensions[col_letter].width = width

    ws.freeze_panes = "A2"

    output_buffer = io.BytesIO()
    wb.save(output_buffer)
    output_buffer.seek(0)
    return output_buffer


def _build_zip(selected_reports):
    """Bundles a list of (source_label, row_count, buffer) tuples into a
    single in-memory ZIP for bulk download."""
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for source_label, _row_count, buffer in selected_reports:
            file_name = f"CareLinkExpress-Data-{sanitize_filename_part(source_label)}.xlsx"
            buffer.seek(0)
            zf.writestr(file_name, buffer.read())
            buffer.seek(0)
    zip_buffer.seek(0)
    return zip_buffer


_SOURCE_ICON_SVG = (
    '<svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" '
    'stroke-width="2" stroke-linecap="round" stroke-linejoin="round">'
    '<path d="M3 21h18"/><path d="M5 21V7l7-4 7 4v14"/>'
    '<path d="M9 9h1"/><path d="M14 9h1"/><path d="M9 13h1"/><path d="M14 13h1"/>'
    '<path d="M9 21v-4h6v4"/></svg>'
)


def _reset_selection_state():
    """Clears the permanent selection set, per-row checkbox widget state,
    and Select All — called whenever the set of generated reports changes
    (new upload, new Generate click) so a stale selection from a previous
    file can never carry over."""
    st.session_state["t1_selected"] = set()
    for key in list(st.session_state.keys()):
        if key.startswith("sel_tab1_"):
            del st.session_state[key]
    st.session_state["t1_select_all"] = False


def _sync_select_all_flag():
    reports = st.session_state.get("t1_reports") or []
    selected = st.session_state.get("t1_selected", set())
    st.session_state["t1_select_all"] = bool(reports) and len(selected) == len(reports)


def _on_source_checkbox_change(idx):
    selected = st.session_state.setdefault("t1_selected", set())
    if st.session_state.get(f"sel_tab1_{idx}"):
        selected.add(idx)
    else:
        selected.discard(idx)
    _sync_select_all_flag()


def _on_select_all_toggle():
    reports = st.session_state.get("t1_reports") or []
    selected = st.session_state.setdefault("t1_selected", set())
    if st.session_state.get("t1_select_all"):
        selected.clear()
        selected.update(range(len(reports)))
    else:
        selected.clear()


def render_formatter_tab():
    """Draws the full 'CareLink Express Data' tab UI and wires up processing."""
    st.markdown(
        """
    <div class="section-heading">Rendered Medicines Formatter</div>
    <div class="section-subtext">Upload a raw medicine report — column layout, naming, and order don't matter. We auto-detect and clean the data, then generate one output file per Patient Source.</div>
    """,
        unsafe_allow_html=True,
    )

    uploaded_file = st.file_uploader(
        "Upload a CareLink Express file",
        type=["xlsx"],
        key="single_fmt",
        label_visibility="visible",
    )

    # Uploading a new file, or removing the current one (the uploader's own
    # "x"), invalidates any previously generated output so a stale report
    # can never be shown next to the wrong (or no) source file.
    file_signature = (
        f"{getattr(uploaded_file, 'file_id', None) or uploaded_file.name}-{uploaded_file.size}"
        if uploaded_file
        else None
    )
    if st.session_state.get("t1_file_signature") != file_signature:
        st.session_state["t1_file_signature"] = file_signature
        st.session_state["t1_processed"] = False
        st.session_state["t1_reports"] = None
        _reset_selection_state()

    st.markdown("<br>", unsafe_allow_html=True)

    btn_click = st.button(
        "Generate",
        key="btn_gen_tab1",
        disabled=(uploaded_file is None),
    )

    if uploaded_file and btn_click:
        st.session_state["t1_processed"] = True
        try:
            with st.spinner("Cleaning, formatting, and splitting by Patient Source..."):
                file_bytes = uploaded_file.read()
                df_raw = pd.read_excel(io.BytesIO(file_bytes))
                st.session_state["t1_reports"] = process_rendered_medicines(df_raw)
            _reset_selection_state()
        except Exception:
            st.session_state["t1_reports"] = None
            st.error(
                "Unable to process this file. Please verify that the file "
                "format and contents are correct."
            )

    reports = st.session_state.get("t1_reports")
    if not (st.session_state.get("t1_processed") and reports):
        return

    n = len(reports)
    st.markdown(
        f"""
    <div class="success-banner">
        Generated {n} output file{'s' if n != 1 else ''} — one per Patient Source.
    </div>
    """,
        unsafe_allow_html=True,
    )

    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown('<div class="card-title" style="margin-bottom:8px;">Outputs by Source</div>', unsafe_allow_html=True)

    col_search, col_zip = st.columns([3, 1.6])
    with col_search:
        search_term = st.text_input(
            "Search sources",
            placeholder="Search by source name…",
            key="t1_search",
            label_visibility="collapsed",
        )

    selected_indices = sorted(st.session_state.get("t1_selected", set()))
    with col_zip:
        st.download_button(
            label=f"Download selected ({len(selected_indices)})",
            data=_build_zip([reports[i] for i in selected_indices]) if selected_indices else b"",
            file_name="CareLinkExpress-Selected.zip",
            mime="application/zip",
            key="dl_tab1_selected",
            disabled=(len(selected_indices) == 0),
        )

    st.checkbox("Select All", key="t1_select_all", on_change=_on_select_all_toggle)

    search_lower = (search_term or "").strip().lower()
    visible_indices = [i for i in range(n) if search_lower in reports[i][0].lower()]

    if search_lower and not visible_indices:
        st.markdown(
            '<div class="card-title">No sources match your search.</div>',
            unsafe_allow_html=True,
        )

    if visible_indices:
        hc_check, hc_source, hc_action = st.columns([0.4, 3.1, 1.2])
        with hc_source:
            st.markdown('<div class="source-list-header"><span>SOURCE</span></div>', unsafe_allow_html=True)
        with hc_action:
            st.markdown('<div class="source-list-header"><span>ACTION</span></div>', unsafe_allow_html=True)

    selected_set = st.session_state.get("t1_selected", set())
    for idx in visible_indices:
        source_label, row_count, buffer = reports[idx]
        file_name = f"CareLinkExpress-Data-{sanitize_filename_part(source_label)}.xlsx"
        col_check, col_label, col_btn = st.columns([0.4, 3.1, 1.2])
        with col_check:
            # Force-sync from the permanent selection set on every render.
            # A checkbox's own widget state is NOT reliably preserved by
            # Streamlit across a run where it wasn't drawn at all (e.g.
            # hidden by the search filter), so this set is the real source
            # of truth, not the checkbox's key.
            st.session_state[f"sel_tab1_{idx}"] = idx in selected_set
            st.checkbox(
                "Select",
                key=f"sel_tab1_{idx}",
                label_visibility="collapsed",
                on_change=_on_source_checkbox_change,
                args=(idx,),
            )
        with col_label:
            st.markdown(
                f'<div class="source-row"><span class="source-row-name">{_SOURCE_ICON_SVG} {source_label} '
                f'<span class="source-row-count">({row_count} row{"s" if row_count != 1 else ""})</span></span></div>',
                unsafe_allow_html=True,
            )
        with col_btn:
            st.download_button(
                label="Download",
                data=buffer,
                file_name=file_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"dl_tab1_{idx}",
            )