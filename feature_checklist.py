"""
FEATURE 3: Patient Follow-Up Checklist.

Upload the same two files as the Coordinator merge (Rendered Medicines +
Registered Patients), get back a single hand-markable checklist covering
EVERY patient — both those who already received medicine and those still
awaiting follow-up — with columns:

    1st Contact, Consult, 2nd Contact, Full Name, Cellphone Number,
    Full Address, Medicines rendered, Patient Source, Notes,
    Prescribed, Packed

Column detection is reused from feature_formatter (for the Rendered
Medicines file) and feature_merger (for the Registered Patients lookup)
rather than re-implemented, so fuzzy-matching behavior stays identical
across every feature in the app.
"""
import io

import openpyxl
import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from feature_formatter import _detect_columns, _get
from feature_merger import _build_patient_lookup
from utils import (
    clean_filename,
    clean_pin,
    clean_str,
    format_contact_number,
    read_data_file,
    split_name_fallback,
)

HEADERS = [
    "1st Contact",
    "Consult",
    "2nd Contact",
    "Full Name",
    "Cellphone Number",
    "Full Address",
    "Medicines rendered",
    "Patient Source",
    "Notes",
    "Prescribed",
    "Packed",
]

# Assumption: no explicit dropdown values were given, so these are a
# reasonable default follow-up-call vocabulary. Easy to change — see
# CONSULT_OPTIONS below.
CONSULT_OPTIONS = ["Pending", "Scheduled", "Completed", "No Show"]

CHECKBOX_COLS = {"1st Contact", "2nd Contact", "Prescribed", "Packed"}

COLUMN_WIDTHS = {
    "1st Contact": 11,
    "Consult": 14,
    "2nd Contact": 11,
    "Full Name": 26,
    "Cellphone Number": 15,
    "Full Address": 30,
    "Medicines rendered": 32,
    "Patient Source": 20,
    "Notes": 24,
    "Prescribed": 11,
    "Packed": 10,
}


def _full_name(last, first, middle):
    return " ".join(p for p in [first, middle, last] if p)


def _build_rendered_rows(df_med):
    """One row per PATIENT (not per medicine) — medicines they received
    are aggregated into a single comma-separated cell."""
    cols = _detect_columns(df_med)
    by_pin = {}
    order = []

    for _, row in df_med.iterrows():
        pin = clean_pin(_get(row, cols["pin"])) if cols["pin"] else ""

        last_val = clean_str(_get(row, cols["last"])) if cols["last"] else ""
        first_val = clean_str(_get(row, cols["first"])) if cols["first"] else ""
        middle_val = clean_str(_get(row, cols["middle"])) if cols["middle"] else ""
        raw_name = clean_str(_get(row, cols["patient_name"])) if cols["patient_name"] else ""

        if last_val and first_val:
            full_name = raw_name or _full_name(last_val, first_val, middle_val)
        elif raw_name:
            full_name = raw_name
        else:
            l, f, m = split_name_fallback(raw_name)
            full_name = _full_name(l, f, m)

        key = pin or f"NO_PIN::{full_name.lower()}"
        if key not in by_pin:
            by_pin[key] = {
                "full_name": full_name,
                "phone": format_contact_number(_get(row, cols["phone"])) if cols["phone"] else "",
                "address": clean_str(_get(row, cols["address"])) if cols["address"] else "",
                "source": clean_str(_get(row, cols["source"])) if cols["source"] else "",
                "medicines": [],
            }
            order.append(key)

        medicine = clean_str(_get(row, cols["medicine"])) if cols["medicine"] else ""
        if medicine and medicine not in by_pin[key]["medicines"]:
            by_pin[key]["medicines"].append(medicine)
        if not by_pin[key]["source"] and cols["source"]:
            by_pin[key]["source"] = clean_str(_get(row, cols["source"]))

    return {
        pin: {
            "Full Name": info["full_name"],
            "Cellphone Number": info["phone"],
            "Full Address": info["address"],
            "Medicines rendered": ", ".join(info["medicines"]),
            "Patient Source": info["source"],
        }
        for pin, info in ((k, by_pin[k]) for k in order)
    }, cols


def build_checklist(df_med, df_patient):
    rendered_by_pin, med_cols = _build_rendered_rows(df_med)

    patient_lookup, patient_meta = _build_patient_lookup(df_patient)
    if patient_lookup is None:
        st.error(
            "Could not find a 'Patient PIN' column in the Registered "
            f"Patients file. Found columns: {list(df_patient.columns)}"
        )
        return None, None

    rows = []
    seen_pins = set()

    for pin, info in rendered_by_pin.items():
        rows.append(dict(info))
        if not pin.startswith("NO_PIN::"):
            seen_pins.add(pin)

    for pin, info in patient_lookup.items():
        if pin in seen_pins:
            continue
        rows.append(
            {
                "Full Name": _full_name(info["Last Name"], info["First Name"], info["Middle Name"]),
                "Cellphone Number": info["Contact Number"],
                "Full Address": info["Full Address"],
                "Medicines rendered": "",
                "Patient Source": "",
            }
        )

    rows.sort(key=lambda r: r["Full Name"].strip().lower())

    with st.expander("Detected column mapping (click to verify)"):
        st.markdown(
            f"""
**Rendered Medicines file**
- Patient PIN → {f"`{med_cols['pin']}`" if med_cols['pin'] else "not found — left blank"}
- Name → {f"`{med_cols['last']}` + `{med_cols['first']}`" if med_cols['last'] and med_cols['first'] else (f"split from `{med_cols['patient_name']}`" if med_cols['patient_name'] else "not found — left blank")}
- Contact Number → {f"`{med_cols['phone']}`" if med_cols['phone'] else "not found — left blank"}
- Full Address → {f"`{med_cols['address']}`" if med_cols['address'] else "not found — left blank"}
- Medicine → {f"`{med_cols['medicine']}`" if med_cols['medicine'] else "not found — left blank"}
- Patient Source → {f"`{med_cols['source']}`" if med_cols['source'] else "not found — left blank"}

**Registered Patients file**
- Patient PIN → `{patient_meta['pin_col']}`
- Name → {patient_meta['name_source']}
- Contact Number → {f"`{patient_meta['phone_col']}`" if patient_meta['phone_col'] else "not found — left blank"}
- Full Address → {f"`{patient_meta['address_col']}`" if patient_meta['address_col'] else "not found — left blank"}

**Result:** {len(rows)} total patients — {len(seen_pins)} already rendered, {len(rows) - len(seen_pins)} awaiting follow-up.
            """
        )

    return _build_workbook(rows), len(rows)


def _build_workbook(rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Patient Checklist"
    ws.views.sheetView[0].showGridLines = True

    header_font = Font(name="Calibri", size=11, bold=True)
    header_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="medium"),
    )
    data_font = Font(name="Calibri", size=10)
    thin_border = Border(
        left=Side(style="thin", color="000000"), right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"), bottom=Side(style="thin", color="000000"),
    )
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 24
    for col_idx, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.alignment = align_center
        cell.border = header_border

    for row_idx, row in enumerate(rows, start=2):
        ws.row_dimensions[row_idx].height = 20
        for col_idx, header in enumerate(HEADERS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.border = thin_border
            if header in CHECKBOX_COLS:
                cell.value = ""
                cell.alignment = align_center
            elif header == "Consult":
                cell.value = ""
                cell.alignment = align_center
            elif header == "Cellphone Number":
                cell.value = row.get(header, "")
                cell.number_format = "@"
                cell.alignment = align_center
            else:
                cell.value = row.get(header, "")
                cell.alignment = align_left

    if rows:
        last_row = len(rows) + 1
        dv = DataValidation(
            type="list",
            formula1='"{}"'.format(",".join(CONSULT_OPTIONS)),
            allow_blank=True,
            showDropDown=False,
        )
        ws.add_data_validation(dv)
        dv.add(f"B2:B{last_row}")

    for col_idx, header in enumerate(HEADERS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = COLUMN_WIDTHS.get(header, 16)

    ws.freeze_panes = "A2"

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def render_checklist_tab():
    """Draws the full 'CareLink Checklist Data' tab UI and wires up processing."""
    st.markdown(
        """
    <div class="section-heading">Patient Follow-Up Checklist</div>
    <div class="section-subtext">Upload Rendered Medicines and Registered Patients files to generate one hand-markable checklist covering every patient — both already served and still awaiting follow-up.</div>
    """,
        unsafe_allow_html=True,
    )

    col1, col2 = st.columns(2)

    with col1:
        st.markdown(
            """
        <div class="uploader-card">
            <div class="uploader-title uploader-accent-blue">01 — RENDERED MEDICINES FILE</div>
            <div class="uploader-sub">Upload Rendered Medicines (.xlsx or .csv)</div>
        </div>
        """,
            unsafe_allow_html=True,
        )
        file_med = st.file_uploader(
            "Upload Rendered Medicines",
            type=["xlsx", "csv"],
            key="tab3_med",
            label_visibility="collapsed",
        )

    with col2:
        st.markdown(
            """
        <div class="uploader-card">
            <div class="uploader-title uploader-accent-green">02 — REGISTERED PATIENTS FILE</div>
            <div class="uploader-sub">Upload Registered Patients (.xlsx or .csv)</div>
        </div>
        """,
            unsafe_allow_html=True,
        )
        file_patient = st.file_uploader(
            "Upload Registered Patients",
            type=["xlsx", "csv"],
            key="tab3_pat",
            label_visibility="collapsed",
        )

    def _sig(f):
        return f"{getattr(f, 'file_id', None) or f.name}-{f.size}" if f else None

    file_signature = (_sig(file_med), _sig(file_patient))
    if st.session_state.get("t3_file_signature") != file_signature:
        st.session_state["t3_file_signature"] = file_signature
        st.session_state["t3_processed"] = False
        st.session_state["t3_buffer"] = None

    st.markdown("<br>", unsafe_allow_html=True)

    st.markdown(
        '<div class="card-title" style="margin-left: 2px;">Output File Name:</div>',
        unsafe_allow_html=True,
    )

    c_name, c_action = st.columns([3.2, 1.2])
    default_out = "CareLinkChecklist-Data-Source.xlsx"

    with c_name:
        output_file_name_input = st.text_input(
            "Output File Name",
            value=default_out,
            key="filename_tab3",
            label_visibility="collapsed",
        )

    with c_action:
        both_uploaded = (file_med is not None) and (file_patient is not None)
        btn_click = st.button("Generate", key="btn_gen_tab3", disabled=not both_uploaded)

    if not both_uploaded:
        missing = []
        if file_med is None:
            missing.append("Rendered Medicines")
        if file_patient is None:
            missing.append("Registered Patients")
        st.markdown(
            f'<div class="card-title" style="color:#94A3B8;margin-top:-4px;">'
            f'Please upload {" and ".join(missing)} before generating the checklist.</div>',
            unsafe_allow_html=True,
        )

    if both_uploaded and btn_click:
        st.session_state["t3_processed"] = True
        try:
            with st.spinner("Combining both files into one checklist..."):
                df_med = read_data_file(file_med)
                df_patient = read_data_file(file_patient)
                buffer, _count = build_checklist(df_med, df_patient)
                st.session_state["t3_buffer"] = buffer
        except Exception:
            st.session_state["t3_buffer"] = None
            st.error(
                "Something went wrong while processing the file. Please "
                "check the file and try again."
            )

    if st.session_state.get("t3_processed") and st.session_state.get("t3_buffer"):
        out_name = clean_filename(output_file_name_input, default_out)
        st.markdown(
            f"""
        <div class="success-banner">
            <strong>{out_name}</strong> generated — one row per patient, ready to print or work from.
        </div>
        """,
            unsafe_allow_html=True,
        )

        st.download_button(
            label="Download Checklist (.xlsx)",
            data=st.session_state["t3_buffer"],
            file_name=out_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_tab3",
        )