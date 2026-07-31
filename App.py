import io
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import streamlit as st
import streamlit.components.v1 as components

# Page Configuration
st.set_page_config(
    page_title="CareLink Data Formatting Suite",
    page_icon="🧩",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# Custom CSS matching exact screenshot UI/UX design
st.markdown(
    """
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&family=JetBrains+Mono:wght@400;500&display=swap');

html, body, [class*="css"] {
    font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif;
}

/* Background */
.stApp {
    background-color: #EDF2F8 !important;
}

header[data-testid="stHeader"] {
    display: none;
}

.block-container {
    padding-top: 1.8rem !important;
    padding-bottom: 2rem !important;
    max-width: 980px !important;
}

/* App Header */
.app-header {
    display: flex;
    align-items: center;
    gap: 8px;
    font-size: 18px;
    font-weight: 700;
    color: #0F172A;
    margin-bottom: 24px;
}

/* Section Titles */
.hero-tag {
    font-size: 11px;
    font-weight: 700;
    letter-spacing: 0.08em;
    text-transform: uppercase;
    margin-bottom: 4px;
}
.hero-tag-blue { color: #0284C7; }
.hero-tag-green { color: #1E3A5F; }

.hero-title {
    font-size: 32px;
    font-weight: 800;
    color: #0F172A;
    margin-bottom: 6px;
    line-height: 1.1;
    letter-spacing: -0.02em;
}

.hero-subtitle {
    font-size: 14px;
    color: #64748B;
    line-height: 1.5;
    margin-bottom: 24px;
}

/* Card Container Base */
.card-box {
    background-color: #FFFFFF;
    border: 1px solid #E2E8F0;
    border-radius: 12px;
    padding: 20px;
    margin-bottom: 16px;
    box-shadow: 0 1px 2px rgba(0,0,0,0.02);
}

.card-title {
    font-size: 13.5px;
    color: #334155;
    font-weight: 500;
    margin-bottom: 12px;
}

/* Instruction Tags Grid */
.columns-grid {
    display: flex;
    flex-wrap: wrap;
    gap: 6px;
    margin-top: 12px;
}
.column-tag {
    background-color: #F8FAFC;
    color: #475569;
    font-size: 12px;
    font-weight: 500;
    padding: 5px 12px;
    border-radius: 6px;
    border: 1px solid #E2E8F0;
}

/* File Upload Display Box */
.uploader-card {
    background-color: #FFFFFF;
    border: 1px solid #E2E8F0;
    border-radius: 12px;
    padding: 16px 20px;
    margin-bottom: 16px;
}

.uploader-title {
    font-size: 11px;
    font-weight: 700;
    letter-spacing: 0.06em;
    text-transform: uppercase;
    margin-bottom: 4px;
}

.uploader-sub {
    font-size: 13px;
    color: #475569;
    margin-bottom: 12px;
}

.fake-upload-slot {
    background-color: #F8FAFC;
    border: 1px solid #E2E8F0;
    border-radius: 8px;
    padding: 10px 14px;
    display: flex;
    align-items: center;
    gap: 12px;
    font-family: 'JetBrains Mono', monospace;
    font-size: 13px;
}

/* Monospace Input Styling */
div[data-baseweb="input"] input {
    font-family: 'JetBrains Mono', monospace !important;
    font-size: 13px !important;
    color: #1E293B !important;
    background-color: #F8FAFC !important;
    border-radius: 8px !important;
    border: 1px solid #E2E8F0 !important;
    padding-left: 14px !important;
}

/* Status Badges */
.badge-wrapper {
    display: flex;
    align-items: center;
    gap: 8px;
    height: 100%;
    align-items: center;
}

.pill-badge {
    padding: 6px 14px;
    border-radius: 8px;
    font-size: 12px;
    font-weight: 700;
    display: inline-flex;
    align-items: center;
    gap: 6px;
}

.pill-active-blue {
    background-color: #EFF6FF;
    color: #0284C7;
    border: 1px solid #BAE6FD;
}

.pill-active-green {
    background-color: #EEF2FB;
    color: #1E3A5F;
    border: 1px solid #C7D2E8;
}

.pill-inactive {
    background-color: #F1F5F9;
    color: #94A3B8;
    border: 1px solid #E2E8F0;
}

/* Tabs Styling */
.stTabs [data-baseweb="tab-list"] {
    gap: 24px;
    border-bottom: 1px solid #E2E8F0;
    padding-bottom: 0px;
    margin-bottom: 24px;
}

.stTabs [data-baseweb="tab"] {
    height: 38px;
    background-color: transparent;
    font-weight: 600;
    font-size: 14px;
    color: #64748B;
    border: none;
    border-bottom: 2.5px solid transparent;
    padding: 0 4px 8px 4px;
}

.stTabs [aria-selected="true"] {
    background-color: transparent !important;
    color: #0F172A !important;
    border-bottom: 2.5px solid #1E3A5F !important;
}

/* =========================================================
   FILE UPLOADER — restyled to match the CareLink mock exactly:
   compact single row, no drag/drop copy, "Upload" button,
   filename shown as a blue mono link next to the button.
   ========================================================= */

/* Outer container: remove Streamlit's default spacing/border chrome
   so our own card wrapper (rendered just above via markdown) is the
   only visible border. */
div[data-testid="stFileUploader"] {
    width: 100%;
}

div[data-testid="stFileUploader"] > section,
div[data-testid="stFileUploaderDropzone"] {
    background-color: #F8FAFC !important;
    border: 1px solid #E2E8F0 !important;
    border-radius: 8px !important;
    padding: 10px 14px !important;
    display: flex !important;
    flex-direction: row !important;
    align-items: center !important;
    gap: 12px !important;
    min-height: unset !important;
}

/* Hide the big icon + "Drag and drop file here" copy, keep only the
   compact "Limit ..." line (JS below strips the word "Limit "). */
div[data-testid="stFileUploaderDropzoneInstructions"] {
    display: flex !important;
    align-items: center !important;
    gap: 0 !important;
    order: 3;
}
div[data-testid="stFileUploaderDropzoneInstructions"] svg {
    display: none !important;
}
div[data-testid="stFileUploaderDropzoneInstructions"] span {
    display: none !important;
}
div[data-testid="stFileUploaderDropzoneInstructions"] small {
    font-family: 'JetBrains Mono', monospace !important;
    font-size: 11.5px !important;
    color: #94A3B8 !important;
}

/* Browse/Upload button */
div[data-testid="stFileUploaderDropzone"] button {
    background-color: #FFFFFF !important;
    border: 1px solid #CBD5E1 !important;
    border-radius: 8px !important;
    color: #0F172A !important;
    font-weight: 600 !important;
    font-size: 13px !important;
    height: 34px !important;
    padding: 0 14px !important;
    box-shadow: none !important;
    order: 1;
}
div[data-testid="stFileUploaderDropzone"] button:hover {
    background-color: #F8FAFC !important;
    border-color: #94A3B8 !important;
}

/* Uploaded file chip — render as a blue mono filename, no size/thumbnail chrome */
div[data-testid="stFileUploaderFile"] {
    background: transparent !important;
    border: none !important;
    padding: 0 !important;
    order: 2;
    flex: 1 1 auto;
    min-width: 0;
}
div[data-testid="stFileUploaderFile"] > div:first-child {
    display: none !important; /* file type icon */
}
div[data-testid="stFileUploaderFileName"] {
    font-family: 'JetBrains Mono', monospace !important;
    font-size: 13px !important;
    color: #0284C7 !important;
    font-weight: 500 !important;
    white-space: normal !important;
    overflow-wrap: anywhere !important;
}
div[data-testid="stFileUploaderFile"] small {
    display: none !important; /* file size */
}
div[data-testid="stFileUploaderDeleteBtn"] {
    order: 4;
}
div[data-testid="stFileUploaderDeleteBtn"] button svg {
    color: #94A3B8 !important;
}

/* Primary Action Buttons */
div.stButton > button {
    width: 100%;
    height: 42px;
    border-radius: 8px;
    background-color: #1E3A5F !important;
    color: #FFFFFF !important;
    font-size: 14px !important;
    font-weight: 700 !important;
    border: none !important;
    box-shadow: 0 1px 3px rgba(30, 58, 95, 0.3);
    transition: all 0.15s ease;
}

div.stButton > button:hover {
    background-color: #16304C !important;
}

div.stButton > button:disabled {
    background-color: #CBD5E1 !important;
    color: #64748B !important;
    box-shadow: none !important;
    cursor: not-allowed;
}

/* Banner */
.success-banner {
    background-color: #EFF6FF;
    border: 1px solid #BAE6FD;
    border-radius: 8px;
    padding: 12px 16px;
    color: #1E3A5F;
    font-size: 13px;
    font-weight: 500;
    margin-top: 16px;
}
</style>
""",
    unsafe_allow_html=True,
)


def inject_uploader_tweaks():
    """
    Streamlit's file_uploader widget has no prop to change its copy
    ("Browse files" -> "Upload", "Limit 200MB..." -> "200MB...").
    This runs a tiny script in the parent document (same-origin iframe
    trick) to rewrite that text so it matches the mock, and re-applies
    itself on every rerun via a MutationObserver since Streamlit
    re-renders the DOM on each interaction.
    """
    components.html(
        """
        <script>
        (function() {
            const doc = window.parent.document;

            function tweak() {
                // "Browse files" -> "Upload"
                doc.querySelectorAll('div[data-testid="stFileUploaderDropzone"] button').forEach(btn => {
                    const label = btn.querySelector('div, span, p') || btn;
                    if (label && label.textContent.trim().toLowerCase().includes('browse')) {
                        label.textContent = '↑  Upload';
                    }
                });
                // "Limit 200MB per file • XLSX" -> "200MB per file • XLSX"
                doc.querySelectorAll('div[data-testid="stFileUploaderDropzoneInstructions"] small').forEach(el => {
                    if (el.textContent.startsWith('Limit ')) {
                        el.textContent = el.textContent.replace('Limit ', '');
                    }
                });
            }

            tweak();
            const observer = new MutationObserver(tweak);
            observer.observe(doc.body, { childList: true, subtree: true });
        })();
        </script>
        """,
        height=0,
        width=0,
    )


inject_uploader_tweaks()


def read_data_file(uploaded_file):
    """Helper to read CSV or XLSX file."""
    if uploaded_file.name.endswith(".csv"):
        return pd.read_csv(uploaded_file)
    else:
        return pd.read_excel(uploaded_file)


def format_contact_number(val):
    """Formats phone number to 11 digits (09xxxxxxxxx)."""
    if pd.isna(val) or str(val).strip() == "":
        return ""
    val_str = str(val).split(".")[0].strip()
    if len(val_str) == 10 and val_str.startswith("9"):
        return "0" + val_str
    return val_str


def clean_filename(user_input, default_name):
    """Ensures file name ends with .xlsx and is not empty."""
    name = user_input.strip() if user_input else ""
    if not name or name == ".xlsx":
        return default_name
    if not name.lower().endswith(".xlsx"):
        name += ".xlsx"
    return name


# =========================================================
# FEATURE 1: Rendered Medicines Report Formatter
# =========================================================
def process_rendered_medicines(df):
    for date_col in ["Consultation Date", "Rendered Date"]:
        if date_col in df.columns:
            df[date_col] = (
                pd.to_datetime(df[date_col], errors="coerce")
                .dt.strftime("%m/%d/%Y")
                .fillna("")
            )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rendered Medicines"
    ws.views.sheetView[0].showGridLines = True

    merge_cols = [
        "Patient Name",
        "Last Name",
        "First Name",
        "Middle Name",
        "Patient PIN",
        "Patient Source",
        "Consultation Date",
        "Rendered Date",
        "End Visit By",
        "ICD10 Code",
        "ICD10 Description",
        "Yakap Status",
        "Pharmacy",
        "Contact Number",
        "Address",
        "Notes",
    ]

    headers = list(df.columns)

    header_fill = PatternFill(
        start_color="1F497D", end_color="1F497D", fill_type="solid"
    )
    header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    data_font = Font(name="Segoe UI", size=9.5)
    total_font = Font(name="Segoe UI", size=10, bold=True)

    fill_even = PatternFill(
        start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
    )
    fill_odd = PatternFill(
        start_color="F9FAFB", end_color="F9FAFB", fill_type="solid"
    )

    thin_border_side = Side(border_style="thin", color="D9D9D9")
    thin_border = Border(
        left=thin_border_side,
        right=thin_border_side,
        top=thin_border_side,
        bottom=thin_border_side,
    )
    top_thick_bottom_double = Border(
        top=Side(border_style="thin", color="1F497D"),
        bottom=Side(border_style="double", color="1F497D"),
    )

    align_header = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center")
    align_center = Alignment(horizontal="center", vertical="center")
    align_center_top = Alignment(
        horizontal="center", vertical="top", wrap_text=True
    )
    align_left_top = Alignment(
        horizontal="left", vertical="top", wrap_text=True
    )

    ws.row_dimensions[1].height = 28
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_header
        cell.border = thin_border

    patient_blocks = []
    current_pin = None
    block_start = 2

    for row_idx, row in df.iterrows():
        pin = (
            str(row["Patient PIN"])
            if ("Patient PIN" in df.columns and pd.notna(row["Patient PIN"]))
            else f"NO_PIN_{row_idx}"
        )
        excel_row = row_idx + 2
        if current_pin is None:
            current_pin = pin
            block_start = excel_row
        elif pin != current_pin:
            patient_blocks.append((block_start, excel_row - 1))
            current_pin = pin
            block_start = excel_row

    if block_start <= len(df) + 1:
        patient_blocks.append((block_start, len(df) + 1))

    for row_idx, row_data in enumerate(df.values, start=2):
        ws.row_dimensions[row_idx].height = 20
        block_idx = next(
            i for i, (s, e) in enumerate(patient_blocks) if s <= row_idx <= e
        )
        row_fill = fill_even if block_idx % 2 == 0 else fill_odd

        for col_idx, val in enumerate(row_data, start=1):
            col_name = headers[col_idx - 1]
            cell = ws.cell(row=row_idx, column=col_idx)

            cell.value = "" if pd.isna(val) else val
            cell.font = data_font
            cell.fill = row_fill
            cell.border = thin_border

            if col_name in ["Cost", "Price", "Total Cost", "Total Price"]:
                cell.number_format = "₱#,##0.00"
                cell.alignment = align_right
            elif col_name in ["Qty Prescribed", "Qty Dispensed"]:
                cell.number_format = "#,##0"
                cell.alignment = align_right
            elif col_name in [
                "Consultation Date",
                "Rendered Date",
                "Patient PIN",
                "Contact Number",
                "ICD10 Code",
                "Yakap Status",
                "Pharmacy",
            ]:
                cell.alignment = align_center
                if col_name in ["Patient PIN", "Contact Number"]:
                    cell.number_format = "@"
            else:
                cell.alignment = align_left

    for col_name in merge_cols:
        if col_name in headers:
            col_idx = headers.index(col_name) + 1
            for start_row, end_row in patient_blocks:
                if start_row < end_row:
                    sub_vals = [
                        ws.cell(r, col_idx).value
                        for r in range(start_row, end_row + 1)
                    ]
                    if len(set(sub_vals)) == 1:
                        ws.merge_cells(
                            start_row=start_row,
                            start_column=col_idx,
                            end_row=end_row,
                            end_column=col_idx,
                        )
                        is_centered = col_name in [
                            "Consultation Date",
                            "Rendered Date",
                            "Patient PIN",
                            "Contact Number",
                            "ICD10 Code",
                            "Yakap Status",
                            "Pharmacy",
                        ]
                        target_align = (
                            align_center_top if is_centered else align_left_top
                        )

                        for r in range(start_row, end_row + 1):
                            c = ws.cell(row=r, column=col_idx)
                            c.alignment = target_align
                            c.border = thin_border

    total_row_idx = len(df) + 2
    ws.row_dimensions[total_row_idx].height = 24

    total_label = ws.cell(row=total_row_idx, column=1, value="Total")
    total_label.font = total_font
    total_label.alignment = Alignment(horizontal="left", vertical="center")

    for col_idx, col_name in enumerate(headers, start=1):
        cell = ws.cell(row=total_row_idx, column=col_idx)
        cell.font = total_font
        cell.border = top_thick_bottom_double
        col_letter = get_column_letter(col_idx)

        if col_name in ["Qty Prescribed", "Qty Dispensed"]:
            cell.value = f"=SUM({col_letter}2:{col_letter}{total_row_idx-1})"
            cell.number_format = "#,##0"
            cell.alignment = align_right
        elif col_name in ["Total Cost", "Total Price"]:
            cell.value = f"=SUM({col_letter}2:{col_letter}{total_row_idx-1})"
            cell.number_format = "₱#,##0.00"
            cell.alignment = align_right

    for col in ws.columns:
        col_idx = col[0].column
        col_name = headers[col_idx - 1]
        max_len = max(
            len(str(cell.value or "")) for cell in col if cell.value is not None
        )
        col_letter = get_column_letter(col_idx)

        if col_name in ["ICD10 Description", "Medicine"]:
            width = min(max(max_len + 3, 25), 45)
        elif col_name in ["Patient Name", "End Visit By"]:
            width = min(max(max_len + 3, 20), 30)
        elif col_name in ["Total Cost", "Total Price", "Cost", "Price"]:
            width = 14
        elif col_name in ["Qty Prescribed", "Qty Dispensed"]:
            width = 15
        else:
            width = min(max(max_len + 4, 12), 25)

        ws.column_dimensions[col_letter].width = width

    ws.freeze_panes = "A2"

    output_buffer = io.BytesIO()
    wb.save(output_buffer)
    output_buffer.seek(0)
    return output_buffer


# =========================================================
# FEATURE 2: Automated Patient & Contact Merger (2 Files)
# =========================================================
def process_and_merge(df_med, df_patient):
    med_pin_col = "Patient PIN" if "Patient PIN" in df_med.columns else None
    patient_pin_col = (
        "Patient PIN" if "Patient PIN" in df_patient.columns else None
    )

    if not med_pin_col or not patient_pin_col:
        st.error(
            "Could not find 'Patient PIN' column in one or both uploaded files."
        )
        return None

    df_med["PIN_JOIN"] = df_med[med_pin_col].astype(str).str.strip()
    df_patient["PIN_JOIN"] = df_patient[patient_pin_col].astype(str).str.strip()

    patient_lookup = df_patient.drop_duplicates(subset=["PIN_JOIN"]).copy()

    phone_col = (
        "Cellphone Number"
        if "Cellphone Number" in patient_lookup.columns
        else (
            "Contact Number"
            if "Contact Number" in patient_lookup.columns
            else None
        )
    )
    address_col = "Address" if "Address" in patient_lookup.columns else None

    cols_to_fetch = ["PIN_JOIN"]
    if phone_col:
        cols_to_fetch.append(phone_col)
    if address_col:
        cols_to_fetch.append(address_col)

    merged = pd.merge(
        df_med, patient_lookup[cols_to_fetch], on="PIN_JOIN", how="left"
    )

    if "Rendered Date" in merged.columns:
        merged["Rendered Date"] = (
            pd.to_datetime(merged["Rendered Date"], errors="coerce")
            .dt.strftime("%m/%d/%Y")
            .fillna("")
        )

    if phone_col and phone_col in merged.columns:
        merged["Contact Number Clean"] = merged[phone_col].apply(
            format_contact_number
        )
    else:
        merged["Contact Number Clean"] = ""

    merged["Address Clean"] = (
        merged[address_col].fillna("") if address_col else ""
    )
    merged["Category Clean"] = (
        merged["Medicine Category"].fillna("")
        if "Medicine Category" in merged.columns
        else ""
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rendered Medicines"
    ws.views.sheetView[0].showGridLines = True

    headers = [
        "No.",
        "Patient Name",
        "Patient PIN",
        "Contact Number",
        "Address",
        "Rendered Date",
        "Medicine",
        "Category",
        "Call",
        "Text",
        "Remarks",
    ]

    header_font = Font(name="Calibri", size=11, bold=True)
    header_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="medium"),
    )
    data_font = Font(name="Calibri", size=10)
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_top_center = Alignment(
        horizontal="center", vertical="top", wrap_text=True
    )
    align_top_left = Alignment(horizontal="left", vertical="top", wrap_text=True)

    ws.row_dimensions[1].height = 24
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.alignment = align_center
        cell.border = header_border

    patient_blocks = []
    current_pin = None
    block_start = 2
    patient_counter = 1

    for row_idx, row in merged.iterrows():
        pin = str(row["PIN_JOIN"])
        excel_row = row_idx + 2
        if current_pin is None:
            current_pin = pin
            block_start = excel_row
        elif pin != current_pin:
            patient_blocks.append((patient_counter, block_start, excel_row - 1))
            patient_counter += 1
            current_pin = pin
            block_start = excel_row

    if block_start <= len(merged) + 1:
        patient_blocks.append(
            (patient_counter, block_start, len(merged) + 1)
        )

    for row_idx, row in merged.iterrows():
        excel_row = row_idx + 2
        ws.row_dimensions[excel_row].height = 20

        ws.cell(row=excel_row, column=1, value="")
        ws.cell(
            row=excel_row,
            column=2,
            value=str(row.get("Patient Name", "") or ""),
        )
        ws.cell(row=excel_row, column=3, value=str(row.get("PIN_JOIN", "")))
        ws.cell(
            row=excel_row,
            column=4,
            value=str(row.get("Contact Number Clean", "")),
        )
        ws.cell(
            row=excel_row, column=5, value=str(row.get("Address Clean", ""))
        )
        ws.cell(
            row=excel_row,
            column=6,
            value=str(row.get("Rendered Date", "") or ""),
        )
        ws.cell(
            row=excel_row, column=7, value=str(row.get("Medicine", "") or "")
        )
        ws.cell(
            row=excel_row,
            column=8,
            value=str(row.get("Category Clean", "") or ""),
        )
        ws.cell(row=excel_row, column=9, value="")
        ws.cell(row=excel_row, column=10, value="")
        ws.cell(row=excel_row, column=11, value="")

        for c_idx in range(1, 12):
            cell = ws.cell(row=excel_row, column=c_idx)
            cell.font = data_font
            cell.border = thin_border

            if c_idx in [3, 4]:
                cell.number_format = "@"

            if c_idx in [1, 3, 4, 6, 8]:
                cell.alignment = align_center
            else:
                cell.alignment = align_left

    for p_num, start_row, end_row in patient_blocks:
        ws.cell(row=start_row, column=1, value=p_num)

        merge_cols = [1, 2, 3, 4, 5]
        for col_idx in merge_cols:
            if start_row < end_row:
                ws.merge_cells(
                    start_row=start_row,
                    end_row=end_row,
                    start_column=col_idx,
                    end_column=col_idx,
                )

            for r in range(start_row, end_row + 1):
                cell = ws.cell(row=r, column=col_idx)
                cell.border = thin_border
                if col_idx in [1, 3, 4]:
                    cell.alignment = align_top_center
                else:
                    cell.alignment = align_top_left

    col_widths = {
        1: 5,
        2: 32,
        3: 16,
        4: 16,
        5: 25,
        6: 15,
        7: 45,
        8: 10,
        9: 8,
        10: 8,
        11: 20,
    }

    for col_idx, w in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    ws.freeze_panes = "A2"

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


# =========================================================
# APP HEADER & TABS NAVIGATION
# =========================================================

st.markdown(
    """
<div class="app-header">
    <svg width="20" height="20" viewBox="0 0 24 24" fill="#0284C7">
        <path d="M3 3h8v8H3V3zm10 0h8v8h-8V3zM3 13h8v8H3v-8zm10 0h8v8h-8v-8z"/>
    </svg>
    CareLink Data Formatting Suite
</div>
""",
    unsafe_allow_html=True,
)

tab1, tab2 = st.tabs(
    ["📄 CareLink Express Data", "⊕ CareLink Coordinator Data"]
)

# ---------------------------------------------------------
# TAB 1: CareLink Express Data
# ---------------------------------------------------------
with tab1:
    st.markdown(
        """
    <div class="hero-tag hero-tag-blue">RENDERED MEDICINES</div>
    <div class="hero-title">Report Formatter</div>
    <div class="hero-subtitle">Convert raw medicine reports into clean, professionally formatted Excel workbooks.</div>
    """,
        unsafe_allow_html=True,
    )

    # Instruction Card
    st.markdown(
        """
    <div class="card-box">
        <div style="display: flex; align-items: center; gap: 8px; font-weight: 700; font-size: 14px; color: #0F172A;">
            <svg width="18" height="18" viewBox="0 0 24 24" fill="#0284C7"><path d="M19 3H5c-1.1 0-2 .9-2 2v14c0 1.1.9 2 2 2h14c1.1 0 2-.9 2-2V5c0-1.1-.9-2-2-2zm0 16H5V5h14v14zM7 7h10v2H7V7zm0 4h10v2H7v-2zm0 4h7v2H7v-2z"/></svg>
            Required Excel (.xlsx) Column Structure
        </div>
        <div style="font-size: 13px; color: #64748B; margin-top: 4px;">
            Ensure your uploaded spreadsheet contains the following standard headers:
        </div>
        <div class="columns-grid">
            <span class="column-tag">Patient Name</span>
            <span class="column-tag">Last Name</span>
            <span class="column-tag">First Name</span>
            <span class="column-tag">Middle Name</span>
            <span class="column-tag">Patient PIN</span>
            <span class="column-tag">Patient Source</span>
            <span class="column-tag">Consultation Date</span>
            <span class="column-tag">Rendered Date</span>
            <span class="column-tag">End Visit By</span>
            <span class="column-tag">ICD10 Code</span>
            <span class="column-tag">ICD10 Description</span>
            <span class="column-tag">Medicine</span>
            <span class="column-tag">Medicine Category</span>
            <span class="column-tag">Qty Prescribed</span>
            <span class="column-tag">Qty Dispensed</span>
            <span class="column-tag">Cost</span>
            <span class="column-tag">Price</span>
            <span class="column-tag">Total Cost</span>
            <span class="column-tag">Total Price</span>
            <span class="column-tag">Contact Number</span>
            <span class="column-tag">Address</span>
            <span class="column-tag">Notes</span>
        </div>
    </div>
    """,
        unsafe_allow_html=True,
    )

    # File Upload Card Container
    with st.container():
        st.markdown(
            '<div class="card-title" style="margin-left: 2px;">Drop your file here</div>',
            unsafe_allow_html=True,
        )
        uploaded_file = st.file_uploader(
            "Drop your file here",
            type=["xlsx"],
            key="single_fmt",
            label_visibility="collapsed",
        )

    st.markdown("<br>", unsafe_allow_html=True)

    # Output Controls Box
    default_out_tab1 = "CareLinkExpress-Data-Source.xlsx"

    with st.container():
        st.markdown(
            '<div class="card-title" style="margin-left: 2px;">Output File Name:</div>',
            unsafe_allow_html=True,
        )

        col_input, col_btn = st.columns([3.5, 1.2])

        with col_input:
            output_file_name_input1 = st.text_input(
                "Output File Name",
                value=default_out_tab1,
                key="filename_tab1",
                label_visibility="collapsed",
            )

        with col_btn:
            btn_click1 = st.button(
                "Format Report →",
                key="btn_gen_tab1",
                disabled=(uploaded_file is None),
            )

    if uploaded_file and btn_click1:
        st.session_state["t1_processed"] = True
        with st.spinner("Formatting workbook..."):
            file_bytes = uploaded_file.read()
            df_raw = pd.read_excel(io.BytesIO(file_bytes))
            st.session_state["t1_buffer"] = process_rendered_medicines(df_raw)

    if st.session_state.get("t1_processed") and st.session_state.get(
        "t1_buffer"
    ):
        out_name1 = clean_filename(output_file_name_input1, default_out_tab1)
        st.markdown(
            f"""
        <div class="success-banner">
            ✓ <strong>{out_name1}</strong> processed successfully!
        </div>
        """,
            unsafe_allow_html=True,
        )

        st.download_button(
            label="📥 Download Formatted Excel Report",
            data=st.session_state["t1_buffer"],
            file_name=out_name1,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_tab1",
        )

# ---------------------------------------------------------
# TAB 2: CareLink Coordinator Data
# ---------------------------------------------------------
with tab2:
    st.markdown(
        """
    <div class="hero-tag hero-tag-green">AUTOMATED LOOKUP</div>
    <div class="hero-title">Merge Contacts & Medicines</div>
    <div class="hero-subtitle">Upload <strong>Rendered Medicines</strong> and <strong>Registered Patients</strong> files to perform the merge.</div>
    """,
        unsafe_allow_html=True,
    )

    col1, col2 = st.columns(2)

    with col1:
        st.markdown(
            """
        <div class="uploader-card">
            <div class="uploader-title hero-tag-blue">01 — RENDERED MEDICINES FILE</div>
            <div class="uploader-sub">Upload Medicines (.xlsx or .csv)</div>
        </div>
        """,
            unsafe_allow_html=True,
        )
        file_med = st.file_uploader(
            "Upload Medicines",
            type=["xlsx", "csv"],
            key="tab2_med",
            label_visibility="collapsed",
        )

    with col2:
        st.markdown(
            """
        <div class="uploader-card">
            <div class="uploader-title hero-tag-green">02 — REGISTERED PATIENTS FILE</div>
            <div class="uploader-sub">Upload Registered Patients (.xlsx or .csv)</div>
        </div>
        """,
            unsafe_allow_html=True,
        )
        file_patient = st.file_uploader(
            "Upload Registered Patients",
            type=["xlsx", "csv"],
            key="tab2_pat",
            label_visibility="collapsed",
        )

    st.markdown("<br>", unsafe_allow_html=True)

    # Bottom Settings Card: Output File Name, Dynamic Badges, Merge Button
    st.markdown(
        '<div class="card-title" style="margin-left: 2px;">Output File Name:</div>',
        unsafe_allow_html=True,
    )

    c_name, c_badges, c_action = st.columns([2.4, 1.1, 1.2])

    with c_name:
        output_file_name_input2 = st.text_input(
            "Output File Name",
            value="CareLinkCoordinator-Data-Source.xlsx",
            key="filename_tab2",
            label_visibility="collapsed",
        )

    with c_badges:
        med_class = (
            "pill-badge pill-active-blue" if file_med else "pill-badge pill-inactive"
        )
        pat_class = (
            "pill-badge pill-active-green"
            if file_patient
            else "pill-badge pill-inactive"
        )

        st.markdown(
            f"""
        <div class="badge-wrapper">
            <span class="{med_class}">● MED</span>
            <span class="{pat_class}">● PAT</span>
        </div>
        """,
            unsafe_allow_html=True,
        )

    with c_action:
        both_uploaded = (file_med is not None) and (file_patient is not None)
        btn_click2 = st.button(
            "Run Merge →", key="btn_gen_tab2", disabled=not both_uploaded
        )

    if both_uploaded and btn_click2:
        st.session_state["t2_processed"] = True
        with st.spinner(
            "Matching Patient PINs and generating tracking sheet..."
        ):
            df_med = read_data_file(file_med)
            df_patient = read_data_file(file_patient)
            st.session_state["t2_buffer"] = process_and_merge(
                df_med, df_patient
            )

    if st.session_state.get("t2_processed") and st.session_state.get(
        "t2_buffer"
    ):
        out_name2 = clean_filename(
            output_file_name_input2, "CareLinkCoordinator-Data-Source.xlsx"
        )
        st.markdown(
            f"""
        <div class="success-banner">
            ✅ <strong>{out_name2}</strong> generated successfully!
        </div>
        """,
            unsafe_allow_html=True,
        )

        st.download_button(
            label="📥 Download Merged Tracking Sheet (.xlsx)",
            data=st.session_state["t2_buffer"],
            file_name=out_name2,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_tab2",
        )