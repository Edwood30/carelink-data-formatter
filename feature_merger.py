"""
FEATURE 2: Automated Patient & Contact Merger (2 files).

Rendered Medicines now carries its own Cellphone Number / Full Address /
split name columns, so those are read straight from that file. The
Registered Patients file is only needed to find patients who have NO
rendered-medicine record at all (Sheet 2).

Output workbook:
  Sheet 1 "Rendered Medicines"        - one row per dispensed medicine,
                                         patient identity/contact columns
                                         merged across each patient's rows.
  Sheet 2 "Registered - Not Rendered" - registered patients whose PIN never
                                         appears in the rendered file.
"""
import io

import openpyxl
import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

from utils import clean_filename, find_column, format_contact_number, read_data_file

SHEET1_HEADERS = [
    "No.",
    "Last Name",
    "First Name",
    "Middle Name",
    "Patient PIN",
    "Contact Number",
    "Full Address",
    "Rendered Date",
    "Medicine",
    "Call",
    "Text",
    "Remarks",
]
SHEET1_MERGE_FIELDS = [
    "Last Name",
    "First Name",
    "Middle Name",
    "Patient PIN",
    "Contact Number",
    "Full Address",
]

SHEET2_HEADERS = [
    "No.",
    "Last Name",
    "First Name",
    "Middle Name",
    "Patient PIN",
    "Contact Number",
    "Full Address",
    "Call",
    "Text",
    "Remarks",
]
SHEET2_MERGE_FIELDS = SHEET1_MERGE_FIELDS  # same field set, blocks are just size 1

CENTER_FIELDS = {"No.", "Patient PIN", "Contact Number", "Rendered Date", "Call", "Text"}

COLUMN_WIDTHS = {
    "No.": 5,
    "Last Name": 18,
    "First Name": 16,
    "Middle Name": 16,
    "Patient PIN": 16,
    "Contact Number": 16,
    "Full Address": 32,
    "Rendered Date": 14,
    "Medicine": 40,
    "Call": 8,
    "Text": 8,
    "Remarks": 22,
}


def _clean_str(val):
    return "" if pd.isna(val) else str(val).strip()


def _clean_pin_value(val):
    s = _clean_str(val)
    if s.endswith(".0"):
        s = s[:-2]
    return s


def _split_name_fallback(name):
    """Best-effort split of a single combined name field into Last/First/Middle."""
    name = _clean_str(name)
    if not name:
        return "", "", ""
    if "," in name:
        last, rest = name.split(",", 1)
        rest_parts = rest.strip().split()
        first = rest_parts[0] if rest_parts else ""
        middle = " ".join(rest_parts[1:]) if len(rest_parts) > 1 else ""
        return last.strip(), first, middle
    parts = name.split()
    if len(parts) == 1:
        return "", parts[0], ""
    first = parts[0]
    last = parts[-1]
    middle = " ".join(parts[1:-1])
    return last, first, middle


def _detect_med_columns(df_med):
    category_col = find_column(df_med, ["Medicine Category", "Category", "Drug Category"])
    return {
        "pin": find_column(df_med, ["Patient PIN", "PIN", "PatientPIN"]),
        "last": find_column(df_med, ["Last Name"]),
        "first": find_column(df_med, ["First Name"]),
        "middle": find_column(df_med, ["Middle Name"]),
        "patient_name": find_column(
            df_med,
            ["Patient Name", "Full Name", "Patient Full Name"],
        ),
        "phone": find_column(
            df_med, ["Cellphone Number", "Contact Number", "Mobile Number", "Phone Number"]
        ),
        "address": find_column(
            df_med, ["Full Address", "Address", "Complete Address", "Home Address"]
        ),
        "rendered_date": find_column(
            df_med, ["Rendered Date", "Date Rendered", "Consultation Date"]
        ),
        "medicine": find_column(
            df_med, ["Medicine", "Medicine Name", "Drug", "Drug Name"], exclude=[category_col]
        ),
    }


def _build_rendered_records(df_med, cols):
    records = []
    for _, row in df_med.iterrows():
        pin = _clean_pin_value(row.get(cols["pin"])) if cols["pin"] else ""

        if cols["last"] and cols["first"]:
            last = _clean_str(row.get(cols["last"]))
            first = _clean_str(row.get(cols["first"]))
            middle = _clean_str(row.get(cols["middle"])) if cols["middle"] else ""
        elif cols["patient_name"]:
            last, first, middle = _split_name_fallback(row.get(cols["patient_name"]))
        else:
            last, first, middle = "", "", ""

        phone = format_contact_number(row.get(cols["phone"])) if cols["phone"] else ""
        address = _clean_str(row.get(cols["address"])) if cols["address"] else ""

        rendered_date_str = ""
        if cols["rendered_date"]:
            parsed = pd.to_datetime(row.get(cols["rendered_date"]), errors="coerce")
            if pd.notna(parsed):
                rendered_date_str = parsed.strftime("%m/%d/%Y")

        medicine = _clean_str(row.get(cols["medicine"])) if cols["medicine"] else ""

        records.append(
            {
                "__pin__": pin,
                "Last Name": last,
                "First Name": first,
                "Middle Name": middle,
                "Patient PIN": pin,
                "Contact Number": phone,
                "Full Address": address,
                "Rendered Date": rendered_date_str,
                "Medicine": medicine,
                "Call": "",
                "Text": "",
                "Remarks": "",
            }
        )
    return records


def _build_patient_lookup(df_patient):
    """Returns (dict of PIN -> patient info, meta dict) or (None, None) if no PIN column."""
    pin_col = find_column(df_patient, ["Patient PIN", "PIN", "PatientPIN"])
    if not pin_col:
        return None, None

    last_col = find_column(df_patient, ["Last Name"])
    first_col = find_column(df_patient, ["First Name"])
    middle_col = find_column(df_patient, ["Middle Name"])
    name_col = find_column(
        df_patient,
        ["Patient Name", "Full Name", "Patient Full Name"],
        exclude=[c for c in (last_col, first_col, middle_col) if c],
    )
    phone_col = find_column(
        df_patient,
        [
            "Cellphone Number",
            "Contact Number",
            "Mobile Number",
            "Phone Number",
            "Cellphone No",
            "Cellphone No.",
            "Contact No",
            "Contact No.",
        ],
    )
    address_col = find_column(
        df_patient, ["Full Address", "Address", "Complete Address", "Home Address"]
    )

    lookup = {}
    for _, row in df_patient.iterrows():
        pin = _clean_pin_value(row.get(pin_col))
        if not pin or pin in lookup:
            continue

        if last_col and first_col:
            last = _clean_str(row.get(last_col))
            first = _clean_str(row.get(first_col))
            middle = _clean_str(row.get(middle_col)) if middle_col else ""
        elif name_col:
            last, first, middle = _split_name_fallback(row.get(name_col))
        else:
            last, first, middle = "", "", ""

        phone = format_contact_number(row.get(phone_col)) if phone_col else ""
        address = _clean_str(row.get(address_col)) if address_col else ""

        lookup[pin] = {
            "Last Name": last,
            "First Name": first,
            "Middle Name": middle,
            "Patient PIN": pin,
            "Contact Number": phone,
            "Full Address": address,
        }

    meta = {
        "pin_col": pin_col,
        "name_source": (
            f"'{last_col}' + '{first_col}' + '{middle_col}'"
            if last_col and first_col
            else (f"'{name_col}' column" if name_col else "not found — left blank")
        ),
        "phone_col": phone_col,
        "address_col": address_col,
    }
    return lookup, meta


def _last_first_key(rec):
    return (
        (rec.get("Last Name") or "").strip().lower(),
        (rec.get("First Name") or "").strip().lower(),
    )


def _sort_records_alphabetically(records):
    """Sorts by Last Name, then First Name — keeping every patient's own rows
    (e.g. multiple medicines) together and in their original relative order."""
    groups = {}
    pin_order = []
    for rec in records:
        pin = rec["__pin__"]
        if pin not in groups:
            groups[pin] = []
            pin_order.append(pin)
        groups[pin].append(rec)

    sorted_pins = sorted(pin_order, key=lambda p: _last_first_key(groups[p][0]))

    sorted_records = []
    for pin in sorted_pins:
        sorted_records.extend(groups[pin])
    return sorted_records


def _apply_column_widths(ws, headers):
    for idx, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = COLUMN_WIDTHS.get(h, 15)


def _write_grouped_sheet(ws, headers, records, merge_field_names):
    """Writes header + data rows, grouping consecutive same-PIN records into
    merged patient blocks (a block of size 1 just renders as a normal row)."""
    ws.views.sheetView[0].showGridLines = True

    header_font = Font(name="Calibri", size=11, bold=True)
    header_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="medium"),
    )
    data_font = Font(name="Calibri", size=10)
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_top_center = Alignment(horizontal="center", vertical="top", wrap_text=True)
    align_top_left = Alignment(horizontal="left", vertical="top", wrap_text=True)

    ws.row_dimensions[1].height = 24
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.alignment = align_center
        cell.border = header_border

    _apply_column_widths(ws, headers)
    ws.freeze_panes = "A2"

    if not records:
        return

    # --- group consecutive rows sharing the same PIN into blocks ---
    blocks = []
    current_pin = None
    block_start = 2
    counter = 1
    for i, rec in enumerate(records):
        excel_row = i + 2
        pin = rec["__pin__"]
        if current_pin is None:
            current_pin = pin
            block_start = excel_row
        elif pin != current_pin:
            blocks.append((counter, block_start, excel_row - 1))
            counter += 1
            current_pin = pin
            block_start = excel_row
    blocks.append((counter, block_start, len(records) + 1))

    # --- write every cell ---
    for i, rec in enumerate(records):
        excel_row = i + 2
        ws.row_dimensions[excel_row].height = 20

        for col_idx, header_name in enumerate(headers, start=1):
            if header_name == "No.":
                continue  # filled in per-block below
            val = rec.get(header_name, "")
            cell = ws.cell(row=excel_row, column=col_idx, value=val)
            cell.font = data_font
            cell.border = thin_border
            if header_name in ("Patient PIN", "Contact Number"):
                cell.number_format = "@"
            cell.alignment = align_center if header_name in CENTER_FIELDS else align_left

        no_cell = ws.cell(row=excel_row, column=1)
        no_cell.font = data_font
        no_cell.border = thin_border
        no_cell.alignment = align_center

    # --- merge patient-level columns across each block ---
    for p_num, start_row, end_row in blocks:
        ws.cell(row=start_row, column=1, value=p_num)

        cols_to_merge = ["No."] + [h for h in merge_field_names if h in headers]
        for header_name in cols_to_merge:
            col_idx = headers.index(header_name) + 1
            if start_row < end_row:
                ws.merge_cells(
                    start_row=start_row, end_row=end_row, start_column=col_idx, end_column=col_idx
                )
            is_centered = header_name in CENTER_FIELDS or header_name == "No."
            target_align = align_top_center if is_centered else align_top_left
            for r in range(start_row, end_row + 1):
                cell = ws.cell(row=r, column=col_idx)
                cell.alignment = target_align
                cell.border = thin_border


def process_and_merge(df_med, df_patient):
    med_cols = _detect_med_columns(df_med)
    if not med_cols["pin"]:
        st.error(
            "Could not find a 'Patient PIN' column in the Rendered Medicines "
            f"file. Found columns: {list(df_med.columns)}"
        )
        return None

    rendered_records = _build_rendered_records(df_med, med_cols)
    rendered_pins = {r["__pin__"] for r in rendered_records if r["__pin__"]}

    patient_lookup, patient_meta = _build_patient_lookup(df_patient)
    if patient_lookup is None:
        st.error(
            "Could not find a 'Patient PIN' column in the Registered Patients "
            f"file. Found columns: {list(df_patient.columns)}"
        )
        return None

    missing_records = []
    for pin, info in patient_lookup.items():
        if pin not in rendered_pins:
            rec = dict(info)
            rec["__pin__"] = pin
            rec["Call"] = ""
            rec["Text"] = ""
            rec["Remarks"] = ""
            missing_records.append(rec)

    rendered_records = _sort_records_alphabetically(rendered_records)
    missing_records = sorted(missing_records, key=_last_first_key)

    with st.expander("Detected column mapping (click to verify)"):
        st.markdown(
            f"""
**Rendered Medicines file**
- Patient PIN → `{med_cols['pin']}`
- Name → {"`" + med_cols['last'] + "` + `" + med_cols['first'] + "` + `" + (med_cols['middle'] or '—') + "`" if med_cols['last'] and med_cols['first'] else (f"split from `{med_cols['patient_name']}`" if med_cols['patient_name'] else "not found — left blank")}
- Contact Number → {f"`{med_cols['phone']}`" if med_cols['phone'] else "not found — left blank"}
- Full Address → {f"`{med_cols['address']}`" if med_cols['address'] else "not found — left blank"}
- Rendered Date → {f"`{med_cols['rendered_date']}`" if med_cols['rendered_date'] else "not found — left blank"}
- Medicine → {f"`{med_cols['medicine']}`" if med_cols['medicine'] else "not found — left blank"}

**Registered Patients file**
- Patient PIN → `{patient_meta['pin_col']}`
- Name → {patient_meta['name_source']}
- Contact Number → {f"`{patient_meta['phone_col']}`" if patient_meta['phone_col'] else "not found — left blank"}
- Full Address → {f"`{patient_meta['address_col']}`" if patient_meta['address_col'] else "not found — left blank"}

**Result:** {len(rendered_records)} rendered-medicine rows across {len(rendered_pins)} patients · {len(missing_records)} registered patient(s) with no rendered-medicine record.
            """
        )

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Rendered Medicines"
    _write_grouped_sheet(ws1, SHEET1_HEADERS, rendered_records, SHEET1_MERGE_FIELDS)

    ws2 = wb.create_sheet("Registered - Not Rendered")
    _write_grouped_sheet(ws2, SHEET2_HEADERS, missing_records, SHEET2_MERGE_FIELDS)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def render_merger_tab():
    """Draws the full 'CareLink Coordinator Data' tab UI and wires up processing."""
    st.markdown(
        """
    <div class="section-heading">Merge Contacts &amp; Medicines</div>
    <div class="section-subtext">Upload <strong>Rendered Medicines</strong> and <strong>Registered Patients</strong> files. Produces a workbook with a Rendered Medicines sheet and a Registered-but-Not-Rendered sheet.</div>
    """,
        unsafe_allow_html=True,
    )

    col1, col2 = st.columns(2)

    with col1:
        st.markdown(
            """
        <div class="uploader-card">
            <div class="uploader-title hero-tag-blue">01 — RENDERED MEDICINES FILE</div>
            <div class="uploader-sub">Upload Medicines (.xlsx or .csv)</div>
        </div>
        """,
            unsafe_allow_html=True,
        )
        file_med = st.file_uploader(
            "Upload Medicines",
            type=["xlsx", "csv"],
            key="tab2_med",
            label_visibility="collapsed",
        )

    with col2:
        st.markdown(
            """
        <div class="uploader-card">
            <div class="uploader-title hero-tag-green">02 — REGISTERED PATIENTS FILE</div>
            <div class="uploader-sub">Upload Registered Patients (.xlsx or .csv)</div>
        </div>
        """,
            unsafe_allow_html=True,
        )
        file_patient = st.file_uploader(
            "Upload Registered Patients",
            type=["xlsx", "csv"],
            key="tab2_pat",
            label_visibility="collapsed",
        )

    # Removing or swapping either file (via the uploader's own "x")
    # invalidates any previously generated output, same as Express.
    def _sig(f):
        return f"{getattr(f, 'file_id', None) or f.name}-{f.size}" if f else None

    file_signature = (_sig(file_med), _sig(file_patient))
    if st.session_state.get("t2_file_signature") != file_signature:
        st.session_state["t2_file_signature"] = file_signature
        st.session_state["t2_processed"] = False
        st.session_state["t2_buffer"] = None

    st.markdown("<br>", unsafe_allow_html=True)

    st.markdown(
        '<div class="card-title" style="margin-left: 2px;">Output File Name:</div>',
        unsafe_allow_html=True,
    )

    c_name, c_badges, c_action = st.columns([2.4, 1.1, 1.2])
    default_out = "CareLinkCoordinator-Data-Source.xlsx"

    with c_name:
        output_file_name_input = st.text_input(
            "Output File Name",
            value=default_out,
            key="filename_tab2",
            label_visibility="collapsed",
        )

    with c_badges:
        med_class = "pill-badge pill-active-blue" if file_med else "pill-badge pill-inactive"
        pat_class = "pill-badge pill-active-green" if file_patient else "pill-badge pill-inactive"

        st.markdown(
            f"""
        <div class="badge-wrapper">
            <span class="{med_class}">● MED</span>
            <span class="{pat_class}">● PAT</span>
        </div>
        """,
            unsafe_allow_html=True,
        )

    with c_action:
        both_uploaded = (file_med is not None) and (file_patient is not None)
        btn_click = st.button(
            "Run Merge", key="btn_gen_tab2", disabled=not both_uploaded
        )

    if both_uploaded and btn_click:
        st.session_state["t2_processed"] = True
        with st.spinner("Matching Patient PINs and generating tracking sheets..."):
            df_med = read_data_file(file_med)
            df_patient = read_data_file(file_patient)
            st.session_state["t2_buffer"] = process_and_merge(df_med, df_patient)

    if st.session_state.get("t2_processed") and st.session_state.get("t2_buffer"):
        out_name = clean_filename(output_file_name_input, default_out)
        st.markdown(
            f"""
        <div class="success-banner">
            <strong>{out_name}</strong> generated — includes a "Rendered Medicines" sheet and a "Registered - Not Rendered" sheet.
        </div>
        """,
            unsafe_allow_html=True,
        )

        st.download_button(
            label="Download Merged Tracking Workbook (.xlsx)",
            data=st.session_state["t2_buffer"],
            file_name=out_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_tab2",
        )